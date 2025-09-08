#!/usr/bin/env python3
"""
RTB2000 Data Export Module
==========================

Advanced data export capabilities:
- Multiple format support (CSV, Excel, HDF5, JSON)
- Waveform data export with metadata
- Measurement results export
- Screenshot and report generation
"""

import numpy as np
import pandas as pd
from PyQt6.QtCore import QObject, pyqtSignal, QThread
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                            QComboBox, QPushButton, QFileDialog, QProgressBar,
                            QTextEdit, QCheckBox, QSpinBox)
from PyQt6.QtGui import QPixmap, QPainter
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging

try:
    import h5py
    HDF5_AVAILABLE = True
except ImportError:
    HDF5_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class DataExporter(QObject):
    """Advanced data export manager"""
    
    # Signals
    export_started = pyqtSignal(str)  # Export format
    export_progress = pyqtSignal(int)  # Progress percentage
    export_completed = pyqtSignal(str)  # File path
    export_failed = pyqtSignal(str)  # Error message
    
    def __init__(self):
        super().__init__()
        self.logger = logging.getLogger(__name__)
        
        # Export settings
        self.export_metadata = True
        self.export_statistics = True
        self.export_measurements = True
        self.compression_enabled = True
        
        # Data to export
        self.waveform_data = {}
        self.measurement_data = []
        self.statistics_data = {}
        self.session_info = {}
        
        self.logger.info("Data Exporter initialized")
    
    def set_waveform_data(self, channel: str, time_data: np.ndarray, 
                         voltage_data: np.ndarray, sample_rate: float,
                         metadata: Dict = None):
        """
        Set waveform data for export
        
        Args:
            channel: Channel name
            time_data: Time array
            voltage_data: Voltage array
            sample_rate: Sampling rate in Hz
            metadata: Additional metadata
        """
        self.waveform_data[channel] = {
            'time': time_data,
            'voltage': voltage_data,
            'sample_rate': sample_rate,
            'metadata': metadata or {},
            'timestamp': datetime.now()
        }
        
        self.logger.debug(f"Waveform data set for {channel}: {len(voltage_data)} points")
    
    def set_measurement_data(self, measurements: List[Dict]):
        """Set measurement data for export"""
        self.measurement_data = measurements
        self.logger.debug(f"Measurement data set: {len(measurements)} measurements")
    
    def set_statistics_data(self, statistics: Dict):
        """Set statistics data for export"""
        self.statistics_data = statistics
        self.logger.debug("Statistics data set")
    
    def set_session_info(self, info: Dict):
        """Set session information for export"""
        self.session_info = info
        self.logger.debug("Session info set")
    
    def export_to_csv(self, file_path: str) -> bool:
        """
        Export data to CSV format
        
        Args:
            file_path: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            self.export_started.emit("CSV")
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header with metadata
                if self.export_metadata:
                    writer.writerow(['# RTB2000 Data Export'])
                    writer.writerow(['# Export Date:', datetime.now().isoformat()])
                    writer.writerow(['# Channels:', ', '.join(self.waveform_data.keys())])
                    writer.writerow([])  # Empty row
                
                # Determine common time base
                if self.waveform_data:
                    # Use the first channel's time data as reference
                    first_channel = list(self.waveform_data.keys())[0]
                    time_data = self.waveform_data[first_channel]['time']
                    
                    # Create header row
                    header = ['Time (s)']
                    for channel in self.waveform_data.keys():
                        header.append(f'{channel} (V)')
                    writer.writerow(header)
                    
                    # Write data rows
                    total_points = len(time_data)
                    for i, time_point in enumerate(time_data):
                        row = [f"{time_point:.9f}"]
                        
                        for channel in self.waveform_data.keys():
                            voltage_data = self.waveform_data[channel]['voltage']
                            if i < len(voltage_data):
                                row.append(f"{voltage_data[i]:.6f}")
                            else:
                                row.append('')
                        
                        writer.writerow(row)
                        
                        # Update progress
                        if i % 1000 == 0:
                            progress = int((i / total_points) * 90)
                            self.export_progress.emit(progress)
                
                # Write measurements if enabled
                if self.export_measurements and self.measurement_data:
                    writer.writerow([])  # Empty row
                    writer.writerow(['# Measurements'])
                    writer.writerow(['Measurement', 'Value', 'Unit', 'Channel', 'Timestamp'])
                    
                    for measurement in self.measurement_data:
                        writer.writerow([
                            measurement.get('name', ''),
                            measurement.get('value', ''),
                            measurement.get('unit', ''),
                            measurement.get('channel', ''),
                            measurement.get('timestamp', '')
                        ])
                
                # Write statistics if enabled
                if self.export_statistics and self.statistics_data:
                    writer.writerow([])  # Empty row
                    writer.writerow(['# Statistics'])
                    writer.writerow(['Parameter', 'Value'])
                    
                    for key, value in self.statistics_data.items():
                        writer.writerow([key, value])
            
            self.export_progress.emit(100)
            self.export_completed.emit(file_path)
            self.logger.info(f"CSV export completed: {file_path}")
            return True
            
        except Exception as e:
            error_msg = f"CSV export failed: {e}"
            self.logger.error(error_msg)
            self.export_failed.emit(error_msg)
            return False
    
    def export_to_excel(self, file_path: str) -> bool:
        """
        Export data to Excel format
        
        Args:
            file_path: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not EXCEL_AVAILABLE:
                raise ImportError("openpyxl not available for Excel export")
            
            self.export_started.emit("Excel")
            
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            
            # Waveform data sheet
            if self.waveform_data:
                ws_waveform = workbook.active
                ws_waveform.title = "Waveform Data"
                
                # Write metadata
                if self.export_metadata:
                    ws_waveform['A1'] = 'RTB2000 Data Export'
                    ws_waveform['A2'] = f'Export Date: {datetime.now().isoformat()}'
                    ws_waveform['A3'] = f'Channels: {", ".join(self.waveform_data.keys())}'
                    
                    start_row = 5
                else:
                    start_row = 1
                
                # Write headers
                headers = ['Time (s)']
                for channel in self.waveform_data.keys():
                    headers.append(f'{channel} (V)')
                
                for col, header in enumerate(headers, 1):
                    ws_waveform.cell(row=start_row, column=col, value=header)
                
                # Write data
                if self.waveform_data:
                    first_channel = list(self.waveform_data.keys())[0]
                    time_data = self.waveform_data[first_channel]['time']
                    
                    for i, time_point in enumerate(time_data):
                        row_num = start_row + 1 + i
                        ws_waveform.cell(row=row_num, column=1, value=time_point)
                        
                        for col, channel in enumerate(self.waveform_data.keys(), 2):
                            voltage_data = self.waveform_data[channel]['voltage']
                            if i < len(voltage_data):
                                ws_waveform.cell(row=row_num, column=col, value=voltage_data[i])
                        
                        # Update progress
                        if i % 1000 == 0:
                            progress = int((i / len(time_data)) * 70)
                            self.export_progress.emit(progress)
            
            # Measurements sheet
            if self.export_measurements and self.measurement_data:
                ws_measurements = workbook.create_sheet("Measurements")
                
                headers = ['Measurement', 'Value', 'Unit', 'Channel', 'Timestamp']
                for col, header in enumerate(headers, 1):
                    ws_measurements.cell(row=1, column=col, value=header)
                
                for row, measurement in enumerate(self.measurement_data, 2):
                    ws_measurements.cell(row=row, column=1, value=measurement.get('name', ''))
                    ws_measurements.cell(row=row, column=2, value=measurement.get('value', ''))
                    ws_measurements.cell(row=row, column=3, value=measurement.get('unit', ''))
                    ws_measurements.cell(row=row, column=4, value=measurement.get('channel', ''))
                    ws_measurements.cell(row=row, column=5, value=measurement.get('timestamp', ''))
            
            # Statistics sheet
            if self.export_statistics and self.statistics_data:
                ws_statistics = workbook.create_sheet("Statistics")
                
                ws_statistics.cell(row=1, column=1, value='Parameter')
                ws_statistics.cell(row=1, column=2, value='Value')
                
                for row, (key, value) in enumerate(self.statistics_data.items(), 2):
                    ws_statistics.cell(row=row, column=1, value=key)
                    ws_statistics.cell(row=row, column=2, value=value)
            
            # Save workbook
            workbook.save(file_path)
            
            self.export_progress.emit(100)
            self.export_completed.emit(file_path)
            self.logger.info(f"Excel export completed: {file_path}")
            return True
            
        except Exception as e:
            error_msg = f"Excel export failed: {e}"
            self.logger.error(error_msg)
            self.export_failed.emit(error_msg)
            return False
    
    def export_to_hdf5(self, file_path: str) -> bool:
        """
        Export data to HDF5 format
        
        Args:
            file_path: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not HDF5_AVAILABLE:
                raise ImportError("h5py not available for HDF5 export")
            
            self.export_started.emit("HDF5")
            
            with h5py.File(file_path, 'w') as hf:
                # Create groups
                waveform_group = hf.create_group('waveforms')
                
                # Export waveform data
                for channel, data in self.waveform_data.items():
                    channel_group = waveform_group.create_group(channel)
                    
                    # Store arrays with compression
                    if self.compression_enabled:
                        channel_group.create_dataset('time', data=data['time'], compression='gzip')
                        channel_group.create_dataset('voltage', data=data['voltage'], compression='gzip')
                    else:
                        channel_group.create_dataset('time', data=data['time'])
                        channel_group.create_dataset('voltage', data=data['voltage'])
                    
                    # Store metadata as attributes
                    channel_group.attrs['sample_rate'] = data['sample_rate']
                    channel_group.attrs['timestamp'] = data['timestamp'].isoformat()
                    
                    for key, value in data['metadata'].items():
                        if isinstance(value, (str, int, float, bool)):
                            channel_group.attrs[key] = value
                
                # Export measurements
                if self.export_measurements and self.measurement_data:
                    measurements_group = hf.create_group('measurements')
                    
                    # Convert measurements to structured arrays
                    if self.measurement_data:
                        measurement_names = [m.get('name', '') for m in self.measurement_data]
                        measurement_values = [m.get('value', 0) for m in self.measurement_data]
                        measurement_units = [m.get('unit', '') for m in self.measurement_data]
                        measurement_channels = [m.get('channel', '') for m in self.measurement_data]
                        
                        measurements_group.create_dataset('names', data=measurement_names)
                        measurements_group.create_dataset('values', data=measurement_values)
                        measurements_group.create_dataset('units', data=measurement_units)
                        measurements_group.create_dataset('channels', data=measurement_channels)
                
                # Export statistics
                if self.export_statistics and self.statistics_data:
                    statistics_group = hf.create_group('statistics')
                    
                    for key, value in self.statistics_data.items():
                        if isinstance(value, (int, float)):
                            statistics_group.create_dataset(key, data=value)
                        else:
                            statistics_group.attrs[key] = str(value)
                
                # Export session info
                if self.session_info:
                    session_group = hf.create_group('session')
                    for key, value in self.session_info.items():
                        if isinstance(value, (str, int, float, bool)):
                            session_group.attrs[key] = value
                
                # Add global metadata
                hf.attrs['export_date'] = datetime.now().isoformat()
                hf.attrs['exporter'] = 'RTB2000 Control Software'
                hf.attrs['version'] = '2.0'
            
            self.export_progress.emit(100)
            self.export_completed.emit(file_path)
            self.logger.info(f"HDF5 export completed: {file_path}")
            return True
            
        except Exception as e:
            error_msg = f"HDF5 export failed: {e}"
            self.logger.error(error_msg)
            self.export_failed.emit(error_msg)
            return False
    
    def export_to_json(self, file_path: str) -> bool:
        """
        Export data to JSON format
        
        Args:
            file_path: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            self.export_started.emit("JSON")
            
            # Prepare data for JSON serialization
            export_data = {
                'export_info': {
                    'date': datetime.now().isoformat(),
                    'exporter': 'RTB2000 Control Software',
                    'version': '2.0'
                }
            }
            
            # Add waveform data
            if self.waveform_data:
                export_data['waveforms'] = {}
                for channel, data in self.waveform_data.items():
                    export_data['waveforms'][channel] = {
                        'time': data['time'].tolist(),
                        'voltage': data['voltage'].tolist(),
                        'sample_rate': data['sample_rate'],
                        'timestamp': data['timestamp'].isoformat(),
                        'metadata': data['metadata']
                    }
            
            # Add measurements
            if self.export_measurements and self.measurement_data:
                export_data['measurements'] = self.measurement_data
            
            # Add statistics
            if self.export_statistics and self.statistics_data:
                export_data['statistics'] = self.statistics_data
            
            # Add session info
            if self.session_info:
                export_data['session'] = self.session_info
            
            # Write JSON file
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False)
            
            self.export_progress.emit(100)
            self.export_completed.emit(file_path)
            self.logger.info(f"JSON export completed: {file_path}")
            return True
            
        except Exception as e:
            error_msg = f"JSON export failed: {e}"
            self.logger.error(error_msg)
            self.export_failed.emit(error_msg)
            return False
    
    def export_screenshot(self, widget, file_path: str, format: str = 'PNG') -> bool:
        """
        Export screenshot of a widget
        
        Args:
            widget: QWidget to capture
            file_path: Output file path
            format: Image format (PNG, JPG, etc.)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            self.export_started.emit(f"Screenshot ({format})")
            
            # Capture widget as pixmap
            pixmap = widget.grab()
            
            # Save pixmap
            success = pixmap.save(file_path, format)
            
            if success:
                self.export_progress.emit(100)
                self.export_completed.emit(file_path)
                self.logger.info(f"Screenshot export completed: {file_path}")
                return True
            else:
                raise Exception("Failed to save screenshot")
            
        except Exception as e:
            error_msg = f"Screenshot export failed: {e}"
            self.logger.error(error_msg)
            self.export_failed.emit(error_msg)
            return False


class ExportWidget(QWidget):
    """Widget for data export controls"""
    
    def __init__(self, data_exporter: DataExporter):
        super().__init__()
        self.data_exporter = data_exporter
        self.setup_ui()
        self.setup_connections()
    
    def setup_ui(self):
        """Setup the export control UI"""
        layout = QVBoxLayout(self)
        
        # Export format selection
        format_layout = QHBoxLayout()
        format_layout.addWidget(QLabel("Format:"))
        
        self.format_combo = QComboBox()
        formats = ['CSV', 'JSON']
        
        if EXCEL_AVAILABLE:
            formats.append('Excel')
        
        if HDF5_AVAILABLE:
            formats.append('HDF5')
        
        self.format_combo.addItems(formats)
        format_layout.addWidget(self.format_combo)
        
        layout.addLayout(format_layout)
        
        # Export options
        options_layout = QVBoxLayout()
        
        self.metadata_check = QCheckBox("Include Metadata")
        self.metadata_check.setChecked(True)
        options_layout.addWidget(self.metadata_check)
        
        self.statistics_check = QCheckBox("Include Statistics")
        self.statistics_check.setChecked(True)
        options_layout.addWidget(self.statistics_check)
        
        self.measurements_check = QCheckBox("Include Measurements")
        self.measurements_check.setChecked(True)
        options_layout.addWidget(self.measurements_check)
        
        self.compression_check = QCheckBox("Enable Compression (HDF5)")
        self.compression_check.setChecked(True)
        options_layout.addWidget(self.compression_check)
        
        layout.addLayout(options_layout)
        
        # Export button
        self.export_btn = QPushButton("Export Data...")
        layout.addWidget(self.export_btn)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Status text
        self.status_text = QTextEdit()
        self.status_text.setMaximumHeight(100)
        self.status_text.setReadOnly(True)
        layout.addWidget(self.status_text)
    
    def setup_connections(self):
        """Setup signal connections"""
        self.export_btn.clicked.connect(self.export_data)
        
        self.data_exporter.export_started.connect(self.on_export_started)
        self.data_exporter.export_progress.connect(self.on_export_progress)
        self.data_exporter.export_completed.connect(self.on_export_completed)
        self.data_exporter.export_failed.connect(self.on_export_failed)
        
        # Update exporter settings when checkboxes change
        self.metadata_check.toggled.connect(
            lambda checked: setattr(self.data_exporter, 'export_metadata', checked)
        )
        self.statistics_check.toggled.connect(
            lambda checked: setattr(self.data_exporter, 'export_statistics', checked)
        )
        self.measurements_check.toggled.connect(
            lambda checked: setattr(self.data_exporter, 'export_measurements', checked)
        )
        self.compression_check.toggled.connect(
            lambda checked: setattr(self.data_exporter, 'compression_enabled', checked)
        )
    
    def export_data(self):
        """Start data export process"""
        try:
            format_name = self.format_combo.currentText()
            
            # File dialog
            file_filters = {
                'CSV': 'CSV files (*.csv)',
                'Excel': 'Excel files (*.xlsx)',
                'HDF5': 'HDF5 files (*.h5)',
                'JSON': 'JSON files (*.json)'
            }
            
            file_filter = file_filters.get(format_name, 'All files (*.*)')
            
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                f"Export {format_name} Data",
                f"rtb2000_data.{format_name.lower()}",
                file_filter
            )
            
            if not file_path:
                return
            
            # Start export
            if format_name == 'CSV':
                self.data_exporter.export_to_csv(file_path)
            elif format_name == 'Excel':
                self.data_exporter.export_to_excel(file_path)
            elif format_name == 'HDF5':
                self.data_exporter.export_to_hdf5(file_path)
            elif format_name == 'JSON':
                self.data_exporter.export_to_json(file_path)
            
        except Exception as e:
            self.status_text.append(f"Export error: {e}")
    
    def on_export_started(self, format_name: str):
        """Handle export started signal"""
        self.export_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_text.append(f"Starting {format_name} export...")
    
    def on_export_progress(self, progress: int):
        """Handle export progress signal"""
        self.progress_bar.setValue(progress)
    
    def on_export_completed(self, file_path: str):
        """Handle export completed signal"""
        self.export_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.status_text.append(f"Export completed: {file_path}")
    
    def on_export_failed(self, error_message: str):
        """Handle export failed signal"""
        self.export_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.status_text.append(f"Export failed: {error_message}")
